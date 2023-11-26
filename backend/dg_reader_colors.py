import json
import openpyxl
import numpy as np
import pandas as pd


def add_extension(element_name, possible_keys, row_dict, json_data):
    found_value = "unknown"
    for key in possible_keys:
        if key in row_dict:
            found_value = row_dict[key]
            break
    if element_name == "Type of mutation":
        found_value = found_value.split()[0]
    elif element_name == "Gene name":
        found_value = found_value.split('_')[0]
    json_data["extension"].append({"url": f"https://{element_name}.com".replace(" ", ""), "valueString": str(found_value)})


def read_dg_excel(path):
    # Load the workbook
    workbook = openpyxl.load_workbook(path, data_only=True)

    jsons = []
    # Iterate over each sheet
    for sheet_name in workbook.sheetnames:
        if not sheet_name.startswith("Variants"):
            continue

        sheet = workbook[sheet_name]
        data = []
        for row in sheet.iter_rows(min_row=2, values_only=True):
            headers = [cell.value for cell in sheet[1]]
            row_dict = dict(zip(headers, row))
            data.append(row_dict)

        for (i, row_dict) in enumerate(data):
            print(row_dict["TMBv4_GRCh38"].split('_')[0], " ", sheet[f"A{i+2}"].fill.start_color.index)

    return jsons


if __name__ == "__main__":
    excel_file_path = 'uploads/DG-14013-23-1A_S1 (paired)_Filtered_variants_DNA.xlsx'
    read_dg_excel(excel_file_path)
    excel_file_path = 'uploads/DG-13217-23-1G_S2 (paired)_Filtered_variants_DNA.xlsx'
    read_dg_excel(excel_file_path)
    excel_file_path = '../EHH23_Icebreaker_AZ23/DG/DG-1-122764-21-j-UMI_S11 (paired)_Filtered_variants_DNA.xlsx'
    read_dg_excel(excel_file_path)
    excel_file_path = '../EHH23_Icebreaker_AZ23/DG/DG-14338-23-plasma-UMI_S10 (paired)_Filtered_variants_DNA.xlsx'
    read_dg_excel(excel_file_path)
